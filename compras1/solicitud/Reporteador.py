## De Reporteador

import csv as aliascsv
import openpyxl

from openpyxl.styles import Font
import  xlwt
import csv
import psycopg2
import io
import math
import json
from django.contrib import messages
from django.http import JsonResponse
from django.urls import reverse_lazy, reverse
from django.utils.encoding import smart_str
from django.views.generic import ListView, CreateView, TemplateView
from reportlab.platypus import *
from reportlab.lib.units import cm
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_JUSTIFY, TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet , ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle
from django.shortcuts import render, get_object_or_404, redirect, HttpResponse, HttpResponseRedirect
from django.conf import settings
from datetime import date, datetime
from reportlab.lib.fonts import addMapping
from reportlab.lib.colors import (
black,
purple,
white,
yellow
)

Story = []
numReporte = 0
nombreReporte = ""
lasColumnas = []
columnas = 0

class PostStoreReportesConsulta(TemplateView):
    print("Entre Reporte1")

    template_name = 'Reportes/ReportesConsulta.html'
    desdeFecha = '2022-01-01'
    hastaFecha = '2022-01-31'


    def stylesheet():
        styles = {
        "default": ParagraphStyle(
            "default",
            fontName="Times-Roman",
            fontSize=10,
            leading=12,
            leftIndent=0,
            rightIndent=0,
            firstLineIndent=0,
            alignment=TA_LEFT,
            spaceBefore=0,
            spaceAfter=0,
            bulletFontName="Times-Roman",
            bulletFontSize=10,
            bulletIndent=0,
            textColor=black,
            backColor=None,
            #wordWrap=None,
            wordWrap='LTR',
            borderWidth=0,
            borderPadding=0,
            borderColor=None,
            borderRadius=None,
            allowWidows=1,
            allowOrphans=0,
            textTransform=None,  # "uppercase" | "lowercase" |                 None
            endDots=None,
            splitLongWords=1,
        )}


        return styles


    def myFirstPage(self, canvas, doc):

        global numReporte
        print("Mi primera Pagina")
        canvas.saveState()
        print("Paso canvas")
        canvas.setFont("Helvetica-Bold",9)
        print("Paso canvas1")

        #print("Numero de reportes = ", numReporte)

        # cabezote
        #logotipo = "C:\\FONDOCM.jpg"
        logotipo = "{% static '/img/medical1.jpg' %}"

        #imagen = Image(logotipo, 0.6 * inch, 0.6 * inch)

        #imagen.hAlign = 'LEFT'

        fecha = date.today()
        format = fecha.strftime('%d / %m / %Y')
        print(fecha)
        print(format)
        canvas.drawImage( "C:\EntornosPython\compras1\compras1\static\img/medical1.jpg", 40, 715, width=50,
                     height=50)
        canvas.drawString(250, 750, "CLINICA MEDICAL")
        canvas.drawString(250, 735, 'NIT: 8305077188')
        canvas.drawString(100, 720, "INFORME: ")
        canvas.drawString(150, 720, nombreReporte)
        canvas.drawString(320, 720, "Fecha:")
        canvas.drawString(360, 720, str(format))
        canvas.drawString(520, 720, "Pág: %d " % (doc.page))


        #canvas.drawString(520, 715, "de %d " % (doc.pageCount))

        # Trae Cabezote
        # Ejemplo Tituolo del reporte conexion a BAse de datos, etc

        texto1 = '_________________________________________________________________________________________________________________'
        canvas.setFont('Helvetica-Bold', 9)
        canvas.drawString(30, 670, texto1)

        sonColumnas = range(0, columnas)
        factor = 30

        for i in sonColumnas:
            print("Tamaño columna = ", len(lasColumnas[i]))
            canvas.drawString(factor, 675, lasColumnas[i])
            factor = factor + 55

        texto1 = '_________________________________________________________________________________________________________________'
        canvas.setFont('Helvetica-Bold', 9)
        canvas.drawString(30, 690, texto1)

        # Fin Cabezote
        pageinfo = "Ejemplo Platypus"
        # canvas.drawString(inch, 0.75 * inch, "Página %d " % (doc.page))
        print ("A restaurar canvas")
        canvas.drawString(200, 20,  "Dirección CALLE 36 SUR 77 33 KENNEDY, BOGOTA")
        canvas.restoreState()

    def myLaterPages(self, canvas, doc):
        print("Entre myLaterPages")

        canvas.saveState()
        canvas.setFont('Helvetica-Bold', 9)
        desdeFecha = '2022-01-01'
        hastaFecha = '2022-01-01'

        print(desdeFecha)
        print(hastaFecha)

        # Inserto codigo desde cabezote
        #logotipo = "C:\\FONDOCM.jpg"
        logotipo = "{% static '/img/medical1.jpg' %}"
        imagen = Image(logotipo, 0.6 * inch, 0.6 * inch)
        imagen.hAlign = 'LEFT'

        fecha = date.today()
        format = fecha.strftime('%d / %m / %Y')
        print(fecha)
        print(format)
        canvas.drawImage(  "C:\EntornosPython\compras1\compras1\static\img/medical1.jpg", 40, 715, width=50,
                 height=50)
        canvas.drawString(250, 750, "CLINICA MEDICAL")
        canvas.drawString(250, 735, 'NIT: 8305077188')
        canvas.drawString(100, 720, "INFORME: ")
        canvas.drawString(150, 720, nombreReporte)
        canvas.drawString(320, 720, "Fecha:")
        canvas.drawString(360, 720, str(format))
        canvas.drawString(520, 720, "Pág: %d " % (doc.page))
        #canvas.drawString(520, 715, "de %d " % (doc.pageCount))



        # Trae Cabezote
        # Ejemplo Tituolo del reporte conexion a BAse de datos, etc



        texto1 = '_________________________________________________________________________________________________________________'
        canvas.setFont('Helvetica-Bold', 9)
        canvas.drawString(30, 670, texto1)

        sonColumnas = range(0, columnas)
        factor = 30

        for i in sonColumnas:
            canvas.drawString(factor, 675, lasColumnas[i])
            factor = factor + 55

        texto1 = '_________________________________________________________________________________________________________________'
        canvas.setFont('Helvetica-Bold', 9)
        canvas.drawString(30, 690, texto1)

        # Fin Cabezote
        pageinfo = "Ejemplo Platypus"
        canvas.drawString(200, 20, "Dirección CALLE 36 SUR 77 33 KENNEDY, BOGOTA")
        # canvas.drawString(inch, 0.75 * inch, "Página %d " % (doc.page))
        canvas.restoreState()

    def get_context_data(self, **kwargs):

        username = self.kwargs['username']
        sedeSeleccionada = self.kwargs['sedeSeleccionada']
        nombreUsuario = self.kwargs['nombreUsuario']
        nombreSede = self.kwargs['nombreSede']
        perfil = self.kwargs['perfil']
        numeroReporte = self.kwargs['numeroReporte']

        context = {}
        print("username = ", username)

        context['Username'] = username
        context['SedeSeleccionada'] = sedeSeleccionada
        context['NombreUsuario'] = nombreUsuario
        context['NombreSede'] = nombreSede
        context['Perfil'] = perfil
        context['NumeroReporte'] = numeroReporte


        ## ESto para reporte 5, que son las Ordenes de Compra excel
        if (numeroReporte == '5'):
            miConexion = psycopg2.connect(host="192.168.0.237", database="bd_solicitudes0", port="5432", user="postgres",
                                      password="BD_m3d1c4l")
            cur = miConexion.cursor()
            #comando = 'SELECT c.id id , c."fechaElab" nombre   FROM public.solicitud_ordenescompra c ORDER BY id'
            #comando = 'select c.id, concat(concat(concat(concat(concat(concat(c."fechaElab",' + "' '),c.contacto), ' '), p.proveedor),' '), usu.nom_usuario) " +  ' from solicitud_ordenescompra c inner join solicitud_proveedores p on (p.id = c.proveedor_id) left join solicitud_usuarios usu on (usu.id = c."responsableCompra_id") ORDER BY id'
            comando = 'select c.id, concat(concat(concat(concat(concat(concat(concat(concat(c."fechaElab",' + "' '),c.contacto), ' '), p.proveedor),' '), usu.nom_usuario), '       '),c.id) " + ' from solicitud_ordenescompra c inner join solicitud_proveedores p on (p.id = c.proveedor_id) left join solicitud_usuarios usu on (usu.id = c."responsableCompra_id") ORDER BY id'

            cur.execute(comando)
            print(comando)

            ordenesCompra = []

            for id,nombre in cur.fetchall():
                ordenesCompra.append({'id': id,"nombre":nombre})

            context['OrdenesCompra'] = ordenesCompra

            print("OrdenesCompra = ", ordenesCompra)

            miConexion.close()

        ## Fin reporte 5

        if (numeroReporte != '5'):
        ## Consigo el listado de coordinadores

            miConexion = psycopg2.connect(host="192.168.0.237", database="bd_solicitudes0", port="5432", user="postgres",
                                      password="BD_m3d1c4l")
            cur = miConexion.cursor()
            comando=""
            if (numeroReporte=='1'):
                print("entre 1")
                comando = "SELECT id,nom_usuario  FROM public.solicitud_usuarios WHERE estadoReg = '" + "A'  ORDER BY nom_usuario"
            if (numeroReporte == '2'):
                comando = "SELECT id,nom_usuario  FROM public.solicitud_usuarios WHERE estadoReg = '" + "A' AND perfil='V' ORDER BY nom_usuario"
            if (numeroReporte == '3'):
                comando = "SELECT id,nom_usuario  FROM public.solicitud_usuarios WHERE estadoReg = '" + "A' AND perfil='A' ORDER BY nom_usuario"

            if (numeroReporte == '4'):
                print ("entre 4")
                comando = "SELECT id,nom_usuario  FROM public.solicitud_usuarios WHERE estadoReg = '" + "A' AND perfil='C' ORDER BY nom_usuario"

            cur.execute(comando)
            print(comando)

            coordinadores = []

            for id, nom_usuario in cur.fetchall():
                coordinadores.append({'id': id, 'nom_usuario': nom_usuario})

            context['Coordinadores'] = coordinadores

            print("coordinadores = ", coordinadores)

            miConexion.close()

        ## fin listado Coordinadores

        return context


    def post(self, request, *args, **kwargs):
        print ("Comenzamos a generar el Informe")
        #  Arrancamos
        # Story = []
        global Story
        global numReporte
        global nombreReporte

        context = {}
        username = request.POST["username"]
        nombreSede = request.POST["nombreSede"]
        nombreUsuario = request.POST["nombreUsuario"]
        sedeSeleccionada = request.POST["sedeSeleccionada"]
        perfil = request.POST["perfil"]

        numeroReporte = request.POST["numeroReporte"]
        tipoArchivo = request.POST["tipoArchivo"]

        print ("tipoArchivo = ", tipoArchivo)




        print ("numeroReporte",numeroReporte)

        ## Aqui meto la rutina para reporte 5 Generar reportes Excel con el retur response y todop

        if (numeroReporte == '5'):
            print ("Entre numeroReporte=",numeroReporte)
            context['Username'] = username
            context['SedeSeleccionada'] = sedeSeleccionada
            context['NombreSede'] = nombreSede
            context['NombreUsuario'] = nombreUsuario
            context['Perfil'] = perfil
            idCompra = request.POST["ordenesCompra"]
            print("ordenCompra", idCompra)

            context['NoOrdenCompra'] = idCompra
            context['SolicitudId'] = 0
            context['success'] = True
            context['message'] = 'Orden de Compra No ' + str(idCompra) + ' Created Successfully!'
            ## Comienzo a preparar la impresion EXCEL  de la Orden de Compra
            print("Voy a abril excel")

            ## Comienza la rutina que crea el archivo Excel

            # Contamos

            context['NoOrdenCompra'] = idCompra
            context['SolicitudId'] = 0
            context['success'] = True
            context['message'] = 'Orden de Compra No ' + str(idCompra) + ' Generada Successfully!'
            ## Comienzo a preparar la impresion EXCEL  de la Orden de Compra
            print("Voy a abril execl")

            # my_wb = openpyxl.Workbook(encoding='ascii')
            my_wb = openpyxl.Workbook()
            print("Voy a abril execl1")
            my_sheet = my_wb.active
            print("Voy a abril execl2")
            fuente1 = Font(name='Century', bold=True, size=11, color='0a0a0a')
            print("Voy a abril execl3")
            fuente2 = Font(name='Century', bold=False, size=11, color='0a0a0a')
            print("Comienzo execl")
            b1 = my_sheet['B1']
            b1.value = "NIT 830.507.718-8"
            e1 = my_sheet['E1']
            e1.value = "FORMATO"
            e1.font = fuente1
            e3 = my_sheet['E3']
            e3.value = "APOYO FINANCIERO COMPRAS"
            e3.font = fuente1
            e5 = my_sheet['E5']
            e5.value = "ORDEN DE COMPRA : " + str(idCompra)

            e5.font = fuente1
            print("pase1")
            j1 = my_sheet['J1']
            j1.value = "Código: FOR-AFI-ORDEN DE COMPRA"

            ## Extraemos los datos Generales de la Orden de compra

            miConexion = psycopg2.connect(host="192.168.0.237", database="bd_solicitudes0", port="5432", user="postgres",     password="BD_m3d1c4l")
            miConexion.set_client_encoding('LATIN1')
            cur = miConexion.cursor()

            comando = 'select c.id id, c."fechaElab" fechaElab, c."fechaRevi" fechaRevi, c."fechaApro" fechaApro, c.contacto contacto,c."entregarEn" entregarEn, c.telefono telefono, c.opciones opciones_id,case when c.opciones =' +  "'A'" + ' then ' + "'Anticipo'" + ' when c.opciones = ' + "'C'" + ' then ' + "'Contra Entrega'" + ' when c.opciones = ' + "'N'" + ' then ' + "'Noventa dias'" + '  end opciones , c."valorBruto" valorBruto, c.descuento descuento, c."valorParcial" valorParcial, c.iva iva, c."valorTotal" valortotal, c.observaciones observaciones ,c.aprobo_id aprobo_id,usu1.nom_usuario nombreAprobo,c."aproboCompraStaff_id" aproboCompraStaff_id, c.area_id area_id,area.area area, c.elaboro_id elaboro_id,usu2.nom_usuario nombreElaboro,c."entragaMercancia_id" entragaMercancia_id,usu3.nom_usuario nombreEntregaMercancia, c.proveedor_id proveedor_id, prov.proveedor proveedor, c."recibeMercancia_id" recibeMercancia_id,usu4.nom_usuario nombrerecibeMercancia,  c."responsableCompra_id" responsableCompra_id, usu5.nom_usuario responsableCompra,  c.revizo_id revizo_id, usu6.nom_usuario nombreRevizo from solicitud_ordenescompra c left join solicitud_usuarios usu1 on (usu1.id = c.aprobo_id) left join solicitud_usuarios usu2 on (usu2.id = c.elaboro_id) left join solicitud_usuarios usu3 on (usu3.id = c."entragaMercancia_id") left join solicitud_usuarios usu4 on (usu4.id = c."recibeMercancia_id") left join solicitud_usuarios usu5 on (usu5.id = c."responsableCompra_id") left join solicitud_usuarios usu6 on (usu6.id = c.revizo_id) inner join solicitud_areas area on (area.id = c.area_id) inner join solicitud_proveedores prov on (prov.id = c.proveedor_id) where c.id=' + idCompra
            print(comando)
            cur.execute("set client_encoding='LATIN1';")
            cur.execute(comando)

            ordenCompra = []

            for id,fechaElab, fechaRevi, fechaApro,contacto, entregarEn,  telefono,opciones_id, opciones , valorBruto, descuento,valorParcial, iva, valortotal, observaciones ,aprobo_id, nombreAprobo,aproboCompraStaff_id, area_id,area, elaboro_id, nombreElaboro,entragaMercancia_id, nombreEntregaMercancia,proveedor_id, proveedor, recibeMercancia_id,nombrerecibeMercancia,responsableCompra_id, responsableCompra,revizo_id,  nombreRevizo   in cur.fetchall():
                ordenCompra.append(
                    {'id': id, 'fechaElab': fechaElab, 'fechaRevi': fechaRevi, 'fechaApro': fechaApro, 'contacto': contacto,
                     'entregarEn':entregarEn, 'telefono':telefono , 'opciones':opciones  , 'valorBruto':valorBruto ,  'descuento':descuento , 'valorParcial':valorParcial , 'iva':iva,
                     'valortotal':valortotal,'observaciones':observaciones,'aprobo_id':aprobo_id, 'nombreAprobo':nombreAprobo, 'aproboCompraStaff_id':aproboCompraStaff_id,
                     'area_id':area_id, 'area':area,'elaboro_id':elaboro_id ,
                     'nombreElaboro':nombreElaboro, 'entragaMercancia_id':entragaMercancia_id,'nombreEntregaMercancia':nombreEntregaMercancia,
                     'proveedor_id':proveedor_id,'proveedor':proveedor, 'recibeMercancia_id':recibeMercancia_id, 'nombrerecibeMercancia':nombrerecibeMercancia,
                     'responsableCompra_id':responsableCompra_id, 'responsableCompra':responsableCompra, 'revizo_id':revizo_id,'nombreRevizo':nombreRevizo
                     })

            miConexion.close()
            print("ordenCompra DATOS MAESTRO : ")
            print(ordenCompra)

            for x in ordenCompra:
                print("X = ", x)
                jsonordenCompra = x

            id = jsonordenCompra['id']
            print("Pasegolo0")
            fechaElab = jsonordenCompra['fechaElab']
            fechaRevi = jsonordenCompra['fechaRevi']
            fechaApro = jsonordenCompra['fechaApro']
            contacto = jsonordenCompra['contacto']
            entregarEn = jsonordenCompra['entregarEn']
            telefono = jsonordenCompra['telefono']
            opciones = jsonordenCompra['opciones']
            valorBruto = jsonordenCompra['valorBruto']
            descuento = jsonordenCompra['descuento']
            valorParcial = jsonordenCompra['valorParcial']
            iva = jsonordenCompra['iva']
            valortotal = jsonordenCompra['valortotal']
            observaciones = jsonordenCompra['observaciones']
            aprobo_id = jsonordenCompra['aprobo_id']
            nombreAprobo = jsonordenCompra['nombreAprobo']
            print("Pasegolo11")
            aproboCompraStaff_id = jsonordenCompra['aproboCompraStaff_id']
            area_id = jsonordenCompra['area_id']
            area = jsonordenCompra['area']
            elaboro_id = jsonordenCompra['elaboro_id']
            nombreElaboro = jsonordenCompra['nombreElaboro']
            entragaMercancia_id = jsonordenCompra['entragaMercancia_id']
            nombreEntregaMercancia = jsonordenCompra['nombreEntregaMercancia']
            proveedor_id = jsonordenCompra['proveedor_id']
            proveedor = jsonordenCompra['proveedor']
            recibeMercancia_id = jsonordenCompra['recibeMercancia_id']
            nombrerecibeMercancia = jsonordenCompra['nombrerecibeMercancia']
            responsableCompra = jsonordenCompra['responsableCompra']
            revizo_id = jsonordenCompra['revizo_id']
            nombreRevizo = jsonordenCompra['nombreRevizo']
            print ("Pasegolo12")







            ## Fin Extraemos los datos Generales de la Orden de compra

            j2 = my_sheet['J2']
            j2.value = "Versión 4"
            j3 = my_sheet['J3']
            j3.value = "Fecha de Elaboración :"
            l3 = my_sheet['L3']
            print ("fechaElab= ",fechaElab)
            l3.value = str(fechaElab)
            l3 = my_sheet['L3']
            l3.value = str(fechaElab)
            j4 = my_sheet['J4']
            j4.value = "Fecha de Revision"
            l4 = my_sheet['L4']
            l4.value = str(fechaRevi)
            j5 = my_sheet['J5']
            j5.value = "Fecha de Aprobacion"
            l5 = my_sheet['L5']
            l5.value = str(fechaApro)
            j6 = my_sheet['J6']
            j6.value = "Pagina"
            j6.font = fuente1
            l6 = my_sheet['L6']
            l6.value = "ESTADO"
            l6.font = fuente1
            print("pase2")
            b7 = my_sheet['B7']
            b7.value = "ELABORO"
            b7.font = fuente1
            c7 = my_sheet['C7']
            c7.value = str(nombreElaboro)
            c7.font = fuente2

            f7 = my_sheet['F7']
            f7.value = "REVIZO"
            f7.font = fuente1
            g7 = my_sheet['G7']
            g7.value = str(nombreRevizo)
            g7.font = fuente2

            j7 = my_sheet['J7']
            j7.value = "APROBO"
            j7.font = fuente1

            k7 = my_sheet['K7']
            k7.value = str(nombreAprobo)
            k7.font = fuente2

            print("pase21")
            e9 = my_sheet['E9']
            e9.value = "DATOS ORDEN DE COMPRA"
            e9.font = fuente1
            b11 = my_sheet['B11']
            b11.value = "FECHA"
            b11.font = fuente2
            d11 = my_sheet['D11']
            d11.value = str(fechaElab)
            print("pase22")
            g11 = my_sheet['G11']
            g11.value = "AREA"
            g11.font = fuente2
            print("pase23")
            h11 = my_sheet['H11']
            h11.value = str(area)

            k11 = my_sheet['K11']
            k11.value = "# Cotizacion:"
            k11.font = fuente1
            l11 = my_sheet['L11']
            l11.value = "Pedido No:"
            l11.font = fuente1
            b12 = my_sheet['B12']
            b12.value = "CONTACTO"
            b12.font = fuente2
            print("pase24")
            d12 = my_sheet['D12']
            d12.value = str(contacto)
            g12 = my_sheet['G12']
            g12.value = "ENTREGAR EN"
            g12.font = fuente2
            h12 = my_sheet['H12']
            h12.value = str(entregarEn)
            b13 = my_sheet['B13']
            b13.value = "TELEFONO"
            b13.font = fuente2
            d13 = my_sheet['D13']
            d13.value = str(telefono)
            b14 = my_sheet['B14']
            b14.value = "          Horario de Recepcion :"
            b14.font = fuente1
            e14 = my_sheet['E14']
            e14.value = "martes y jueves: 7:30 am a 12 pm y de 2:00 pm a 4:00 pm "
            e14.font = fuente1
            e15 = my_sheet['E15']
            e15.value = "DATOS DEL PROVEEDOR"
            e15.font = fuente1

            ## Extraemos los datos del Proveedor

            miConexion = psycopg2.connect(host="192.168.0.237", database="bd_solicitudes0", port="5432", user="postgres",
                                          password="BD_m3d1c4l")
            miConexion.set_client_encoding('LATIN1')
            cur = miConexion.cursor()

            comando = "SELECT prov.proveedor nombre, prov.nit nit, prov.telefono telefono, translate(btrim(prov.direccion::text),'óÓáÁéÉíÍúÚñÑ'::text,'oOaAeEiIuUnN'::text)  direccion, prov.correo correo FROM public.solicitud_proveedores prov WHERE prov.proveedor = '" + str(proveedor) + "'"

            print(comando)
            print("pase26")
            cur.execute("set client_encoding='LATIN1';")
            cur.execute(comando)
            print(comando)

            prov = []

            for nombre, nit, telefono, direccion, correo in cur.fetchall():
                prov.append(
                    {'nombre': nombre, 'nit': nit, 'telefono': telefono, 'direccion': direccion, 'correo': correo})

            miConexion.close()
            print("prov")
            print(prov)

            for x in prov:
                print("X = ", x)
                jsonProv = x

            nombreProveedor = jsonProv['nombre']
            nitProveedor = jsonProv['nit']
            telefonoProveedor = jsonProv['telefono']
            direccionProveedor = jsonProv['direccion']
            correoProveedor = jsonProv['correo']

            print("nombre Proveedor = ", nombreProveedor)
            print("Nit Proveedor = ", nitProveedor)
            print("telefonoProveedor = ", telefonoProveedor)
            print("direccionProveedor = ", direccionProveedor)
            print("correoProveedor = ", correoProveedor)

            ### FIN DATOS DEL PROVEEDOR

            print("Pase 50")
            b16 = my_sheet['B16']
            b16.value = "RAZON SOCIAL"
            b16.font = fuente2
            d16 = my_sheet['D16']
            d16.value = str(nombreProveedor)
            h16 = my_sheet['H16']
            h16.value = "NIT"
            h16.font = fuente1
            print("Pase 51")
            i16 = my_sheet['I16']
            i16.value = str(nitProveedor)
            print("Pase 52")
            k16 = my_sheet['K16']
            k16.value = "TELEFONO:"
            k16.font = fuente1
            print("Pase 53")
            l16 = my_sheet['L16']
            l16.value = str(telefonoProveedor)
            b17 = my_sheet['B17']
            b17.value = "DIRECCION:"
            b17.font = fuente1
            d17 = my_sheet['D17']
            d17.value = str(direccionProveedor)
            print("Pase 54")
            h17 = my_sheet['H17']
            h17.value = "E-MAIL:"
            h17.font = fuente1
            i17 = my_sheet['I17']
            i17.value = str(correoProveedor)
            print("Pase 55")
            b18 = my_sheet['B18']
            b18.value = "ATENCION:"
            b18.font = fuente2
            e20 = my_sheet['E20']
            e20.value = "DETALLE DE LA COMPRA"
            e20.font = fuente1
            # k20 = my_sheet['K20']
            # k20.value = "VALOR BRUTO"
            # k20.font = fuente1
            b21 = my_sheet['B21']
            b21.value = "ITEM"
            b21.font = fuente1
            c21 = my_sheet['C21']
            c21.value = "DESCRIPCION REF"
            c21.font = fuente1
            f21 = my_sheet['F21']
            f21.value = "PRESENT."
            f21.font = fuente1
            g21 = my_sheet['G21']
            g21.value = "IVA"
            g21.font = fuente1
            h21 = my_sheet['H21']
            h21.value = "CANTIDAD"
            h21.font = fuente1
            h22 = my_sheet['H22']
            h22.value = "SOLICITADA"
            h22.font = fuente1
            i22 = my_sheet['I22']
            i22.value = "RECIBIDA"
            i22.font = fuente1
            j21 = my_sheet['J21']
            j21.value = "VALOR UNITARIO"
            j21.font = fuente1
            k21 = my_sheet['K21']
            k21.value = "VALOR TOTAL"
            k21.font = fuente1
            k22 = my_sheet['K22']
            k22.value = "SOLICITADA"
            k22.font = fuente1
            l22 = my_sheet['L22']
            l22.value = "RECIBIDA"
            l22.font = fuente1

            print("Armo pah archivo")
            archivo = 'c:\Apache24\OrdenesCompra\OC_' + str(idCompra) + '.xlsx'
            #archivo = 'c:OC_' + str(idCompra) + '.xlsx'

            print("Archivo =", archivo)

            # Imprimo en un for los valores de los items


            ## Extraemos los items de la Compra


            ## Primero el query de detalle

            miConexion = psycopg2.connect(host="192.168.0.237", database="bd_solicitudes0", port="5432", user="postgres",
                                          password="BD_m3d1c4l")
            miConexion.set_client_encoding('LATIN1')
            cur = miConexion.cursor()

            comando = 'select sol0.item item, art.articulo articulo, des.nombre descripcion, pre.nombre presentacion,sol0.iva iva,sol0."solicitadoOrdenCantidad" solicitadoOrdenCantidad, sol0."recibidoOrdenCantidad" recibidoOrdenCantidad , sol0."valorUnitario" valorUnitario, sol0."solicitadoOrdenValor" solicitadoOrdenValor,  sol0."recibidoOrdenValor" recibidoOrdenValor from solicitud_solicitudesdetalle sol0 inner join solicitud_descripcionCompra des on (des.id = sol0.descripcion_id) inner join solicitud_presentacion pre on (pre.id = sol0.presentacion_id) inner join solicitud_articulos art on (art."codregArticulo" = sol0.producto) where sol0."ordenCompra_id" =' + str(idCompra)  + ' order by sol0.item'
            print(comando)
            cur.execute("set client_encoding='LATIN1';")
            cur.execute(comando)

            ordenCompraDet = []

            for item, articulo, descripcion, presentacion,iva, solicitadoOrdenCantidad, recibidoOrdenCantidad ,valorUnitario, solicitadoOrdenValor, recibidoOrdenValor in cur.fetchall():
                ordenCompraDet.append(
                    {'item': item, 'articulo': articulo, 'descripcion': descripcion, 'presentacion': presentacion,
                     'iva': iva,'solicitadoOrdenCantidad':solicitadoOrdenCantidad, 'recibidoOrdenCantidad':recibidoOrdenCantidad,
                     'valorUnitario':valorUnitario, 'solicitadoOrdenValor':solicitadoOrdenValor, 'recibidoOrdenValor': recibidoOrdenValor})

            miConexion.close()
            print("ordenCompra DATOS DETALLE : ")
            print(ordenCompraDet)

            ##Preguntamos por detalle vacio

            if (ordenCompraDet == []):

                context['message'] = 'Orden de Compra No ' + str(idCompra) + ' Sin detalle!'
                return render(self.request, "Reportes/cabeza.html", context)

            jsonordenCompraDet = {}

            for x in ordenCompraDet:
                print("X = ", x)
                jsonordenCompraDet = x

            item = jsonordenCompraDet['item']
            articulo = jsonordenCompraDet['articulo']
            descripcion = jsonordenCompraDet['descripcion']
            presentacion = jsonordenCompraDet['presentacion']
            iva = jsonordenCompraDet['iva']
            solicitadoOrdenCantidad = jsonordenCompraDet['solicitadoOrdenCantidad']
            recibidoOrdenCantidad = jsonordenCompraDet['recibidoOrdenCantidad']
            valorUnitario = jsonordenCompraDet['valorUnitario']
            solicitadoOrdenValor = jsonordenCompraDet['solicitadoOrdenValor']
            recibidoOrdenValor = jsonordenCompraDet['recibidoOrdenValor']




            ## Fin Query de detakke

            campoItem = 1
            voy = 23
            totalRegistros = len(ordenCompraDet)

            print ("totalRegistros = ",totalRegistros )

            for reg in range(1, totalRegistros + 1):

                llaveb = 'b' + str(voy)
                llaveb1 = 'B' + str(voy)
                llaveb = my_sheet[llaveb1]
                llaveb.value = str(item)

                print("Pase 566")

                # b23 = my_sheet['B23']
                # b23.value = str(data1)
                llavec = 'c' + str(voy)
                llavec1 = 'C' + str(voy)
                llavec = my_sheet[llavec1]
                llavec.value = articulo
                # c23 = my_sheet['C23']
                # c23.value = "AqUi descripcion"

                print("Pase 567")
                llavef = 'f' + str(voy)
                llavef1 = 'F' + str(voy)
                llavef = my_sheet[llavef1]
                llavef.value = presentacion
                # f23 = my_sheet['F23']
                # f23.value = "AqUi presentacion"
                llaveg = 'g' + str(voy)
                llaveg1 = 'G' + str(voy)
                llaveg = my_sheet[llaveg1]
                llaveg.value = str(iva)

                print("Pase 568")
                # g23 = my_sheet['G23']
                # g23.value = str(data2)
                llaveh = 'h' + str(voy)
                llaveh1 = 'H' + str(voy)
                llaveh = my_sheet[llaveh1]
                llaveh.value = str(solicitadoOrdenCantidad)
                # h23 = my_sheet['H23']
                # h23.value = str(data3)
                llavei = 'i' + str(voy)
                llavei1 = 'I' + str(voy)
                llavei = my_sheet[llavei1]
                llavei.value = str(recibidoOrdenCantidad)
                # i23 = my_sheet['I23']
                # i23.value = str(data4)
                print("Pase 569")
                llavej = 'j' + str(voy)
                llavej1 = 'J' + str(voy)
                llavej = my_sheet[llavej1]
                llavej.value = str(valorUnitario)
                # j23 = my_sheet['J23']
                # j23.value = str(data5)
                llavek = 'k' + str(voy)
                llavek1 = 'K' + str(voy)
                llavek = my_sheet[llavek1]
                llavek.value = str(solicitadoOrdenValor)
                # k23 = my_sheet['K23']
                # k23.value = str(data6)
                print("Pase 570")
                llavel = 'l' + str(voy)
                llavel1 = 'L' + str(voy)
                llavel = my_sheet[llavel1]
                llavel.value = str(recibidoOrdenValor)
                # l23 = my_sheet['L23']
                # l23.value = str(data7)
                print("Pase 571")
                voy = voy + 1

                ## Fin detalle Excel

                campoItem = campoItem + 1

            # seguimos con la ultima parte del archivo excel

            voy = voy + 2
            print("Pase 56")

            llaveb = 'b' + str(voy)
            llaveb1 = 'B' + str(voy)
            llaveb = my_sheet[llaveb1]
            llaveb.value = "FORMA DE PAGO"
            llaveb.font = fuente1
            # b27 = my_sheet['B27']
            # b27.value = "FORMA DE PAGO"
            # b27.font = fuente1
            llaveh = 'h' + str(voy)
            llaveh1 = 'H' + str(voy)
            llaveh = my_sheet[llaveh1]
            llaveh.value = "VALOR BRUTO"
            llaveh.font = fuente1
            # h27 = my_sheet['H27']
            # h27.value = "VALOR BRUTO"
            # h27.font = fuente1
            llavel = 'l' + str(voy)
            llavel1 = 'L' + str(voy)
            llavel = my_sheet[llavel1]
            llavel.value = str(valorBruto)
            print("Pase 57")
            # l27 = my_sheet['L27']
            # l27.value = str(form.cleaned_data['valorBruto'])

            voy = voy + 1

            llaveb = 'b' + str(voy)
            llaveb1 = 'B' + str(voy)
            llaveb = my_sheet[llaveb1]
            llaveb.value = "OPCION 1"
            llaveb.font = fuente1
            # b28 = my_sheet['B28']
            # b28.value = "OPCION 1"
            # b28.font = fuente1
            llavec = 'c' + str(voy)
            llavec1 = 'C' + str(voy)
            llavec = my_sheet[llavec1]
            llavec.value = "CONTRA ENTREGA"
            llavec.font = fuente2

            #########################################
            ## AQUI COLOCAR X PARA CONTRA ENTREGA F
            ##########################################

            if (opciones == 'C'):
                llavef = 'f' + str(voy)
                llavef1 = 'F' + str(voy)
                llavef = my_sheet[llavef1]
                # llavef.value = str(form.cleaned_data['opciones'])
                llavef.value = "X"
                llavef.font = fuente2

            # c28 = my_sheet['C28']
            # c28.value = "CONTRA ENTREGA"
            # c28.font = fuente2
            llaveh = 'h' + str(voy)
            llaveh1 = 'H' + str(voy)
            llaveh = my_sheet[llaveh1]
            llaveh.value = "DESCUENTO %"
            llaveh.font = fuente1
            # h28 = my_sheet['H28']
            # h28.value = "DESCUENTO %"
            # h28.font = fuente1
            llavel = 'l' + str(voy)
            llavel1 = 'L' + str(voy)
            llavel = my_sheet[llavel1]
            llavel.value = str(descuento)
            # l28 = my_sheet['L28']
            # l28.value = str(form.cleaned_data['descuento'])

            voy = voy + 1

            llaveb = 'b' + str(voy)
            llaveb1 = 'B' + str(voy)
            llaveb = my_sheet[llaveb1]
            llaveb.value = "OPCION 2"
            llaveb.font = fuente1
            # b29 = my_sheet['B29']
            # b29.value = "OPCION 2"
            # b29.font = fuente1
            llavec = 'c' + str(voy)
            llavec1 = 'C' + str(voy)
            llavec = my_sheet[llavec1]
            llavec.value = "ANTICIPO"
            llavec.font = fuente2

            # c29 = my_sheet['C29']
            # c29.value = "ANTICIPO"
            # c29.font = fuente2
            llavee = 'e' + str(voy)
            llavee1 = 'E' + str(voy)
            llavee = my_sheet[llavee1]
            llavee.value = "50 %"
            llavee.font = fuente2
            # e29 = my_sheet['E29']
            # e29.value = "50 %"
            # e29.font = fuente2

            #########################################
            ## MIRE AQUI ESTA IMPRIMIENDO OPCIONES
            ##########################################

            if (opciones == 'A'):
                llavef = 'f' + str(voy)
                llavef1 = 'F' + str(voy)
                llavef = my_sheet[llavef1]
                # llavef.value = str(form.cleaned_data['opciones'])
                llavef.value = "X"
                llavef.font = fuente2

            # f29 = my_sheet['F29']
            # f29.value = str(form.cleaned_data['opciones'])
            # f29.font = fuente2
            llaveh = 'h' + str(voy)
            llaveh1 = 'H' + str(voy)
            llaveh = my_sheet[llaveh1]
            llaveh.value = "VALOR PARCIAL"
            llaveh.font = fuente1
            # h29 = my_sheet['H29']
            # h29.value = "VALOR PARCIAL"
            # h29.font = fuente1
            llavel = 'l' + str(voy)
            llavel1 = 'L' + str(voy)
            llavel = my_sheet[llavel1]
            llavel.value = str(valorParcial)
            # l29 = my_sheet['L29']
            # l29.value = str(form.cleaned_data['valorParcial'])

            voy = voy + 1

            llavec = 'c' + str(voy)
            llavec1 = 'C' + str(voy)
            llavec = my_sheet[llavec1]
            llavec.value = "CONTRA ENTREGA"
            llavec.font = fuente2
            # c30 = my_sheet['C30']
            # c30.value = "CONTRA ENTREGA"
            # c30.font = fuente2
            llavee = 'e' + str(voy)
            llavee1 = 'E' + str(voy)
            llavee = my_sheet[llavee1]
            llavee.value = "50 %"
            llavee.font = fuente2
            # e30 = my_sheet['E30']
            # e30.value = "50 %"
            # e30.font = fuente2
            llaveh = 'h' + str(voy)
            llaveh1 = 'H' + str(voy)
            llaveh = my_sheet[llaveh1]
            llaveh.value = "IVA"
            llaveh.font = fuente1
            # h30 = my_sheet['H30']
            # h30.value = "IVA"
            # h30.font = fuente1
            llavel = 'l' + str(voy)
            llavel1 = 'L' + str(voy)
            llavel = my_sheet[llavel1]
            llavel.value = str(iva)
            # l30 = my_sheet['L30']
            # l30.value = str(form.cleaned_data['iva'])

            voy = voy + 1

            llaveb = 'b' + str(voy)
            llaveb1 = 'B' + str(voy)
            llaveb = my_sheet[llaveb1]
            llaveb.value = "OPCION 3"
            llaveb.font = fuente1
            # b31 = my_sheet['B31']
            # b31.value = "OPCION 3"
            # b31.font = fuente1
            llavec = 'c' + str(voy)
            llavec1 = 'C' + str(voy)
            llavec = my_sheet[llavec1]
            llavec.value = "NOVENTA (90) DIAS"
            llavec.font = fuente2
            # c31 = my_sheet['C31']
            # c31.value = "NOVENTA (90) DIAS"
            # c31.font = fuente2
            print("Pase 58")

            #########################################
            ## AQUI COLOCAR X PARA 90 DIAS COLUMNA F
            ##########################################

            if (opciones == 'N'):
                llavef = 'f' + str(voy)
                llavef1 = 'F' + str(voy)
                llavef = my_sheet[llavef1]
                # llavef.value = str(form.cleaned_data['opciones'])
                llavef.value = "X"
                llavef.font = fuente2

            voy = voy + 1

            llaveh = 'h' + str(voy)
            llaveh1 = 'H' + str(voy)
            llaveh = my_sheet[llaveh1]
            llaveh.value = "VALOR TOTAL"
            llaveh.font = fuente1
            # h31 = my_sheet['H31']
            # h31.value = "VALOR TOTAL"
            # h31.font = fuente1
            llavel = 'l' + str(voy)
            llavel1 = 'L' + str(voy)
            llavel = my_sheet[llavel1]
            llavel.value = str(valortotal)
            voy = voy + 1
            llaveb = 'b' + str(voy)
            llaveb1 = 'B' + str(voy)
            llaveb = my_sheet[llaveb1]
            llaveb.value = "OPCION 4"
            llaveb.font = fuente1
            # b32 = my_sheet['B32']
            # b32.value = "OPCION 4"
            # b32.font = fuente1
            llaveh = 'h' + str(voy)
            llaveh1 = 'H' + str(voy)
            llaveh = my_sheet[llaveh1]
            llaveh.value = "OBSERVACIONES"
            llaveh.font = fuente1
            # h32 = my_sheet['H32']
            # h32.value = "OBSERVACIONES"
            # h32.font = fuente1

            llavel = 'l' + str(voy)
            llavel1 = 'L' + str(voy)
            llavel = my_sheet[llavel1]
            llavel.value = str(observaciones)
            # l32 = my_sheet['L32']
            # l32.value = str(form.cleaned_data['observaciones'])

            voy = voy + 4

            llaveb = 'b' + str(voy)
            llaveb1 = 'B' + str(voy)
            llaveb = my_sheet[llaveb1]
            llaveb.value = "RESPONSABLE ORDEN DE COMPRA:"
            llaveb.font = fuente1
            # b36 = my_sheet['B36']
            # b36.value = "RESPONSABLE ORDEN DE COMPRA:"
            # b36.font = fuente1
            llaveg = 'g' + str(voy)
            llaveg1 = 'G' + str(voy)
            llaveg = my_sheet[llaveg1]
            llaveg.value = "QUIEN ENTREGA MERCANCIA:"
            llaveg.font = fuente1
            print("Pase 59")
            # g36 = my_sheet['G36']
            # g36.value = "QUIEN ENTREGA MERCANCIA:"
            # g36.font = fuente1
            llavek = 'k' + str(voy)
            llavek1 = 'K' + str(voy)
            llavek = my_sheet[llavek1]
            llavek.value = "QUIEN RECIBE MERCANCIA:"
            llavek.font = fuente1
            # k36 = my_sheet['K36']
            # k36.value = "QUIEN RECIBE MERCANCIA:"
            # k36.font = fuente1

            voy = voy + 2

            llaveb = 'b' + str(voy)
            llaveb1 = 'B' + str(voy)
            llaveb = my_sheet[llaveb1]
            llaveb.value = str(responsableCompra)
            # b38 = my_sheet['B38']
            # b38.value = str(form.cleaned_data['responsableCompra'])
            print("Pase 60")
            llaveg = 'g' + str(voy)
            llaveg1 = 'G' + str(voy)
            llaveg = my_sheet[llaveg1]
            llaveg.value = str(nombreEntregaMercancia)
            # g38 = my_sheet['G38']
            # g38.value = str(form.cleaned_data['entragaMercancia'])
            llavek = 'k' + str(voy)
            llavek1 = 'K' + str(voy)
            llavek = my_sheet[llavek1]
            llavek.value = str(nombrerecibeMercancia)
            # k38 = my_sheet['K38']
            # k38.value = str(form.cleaned_data['recibeMercancia'])

            voy = voy + 5

            print("Pase 61")

            llaveb = 'b' + str(voy)
            llaveb1 = 'B' + str(voy)
            llaveb = my_sheet[llaveb1]
            llaveb.value = "FIRMA Y SELLO"
            llaveb.font = fuente2
            # b43 = my_sheet['B43']
            # b43.value = "FIRMA Y SELLO"
            # b43.font = fuente2
            llaveg = 'g' + str(voy)
            llaveg1 = 'G' + str(voy)
            llaveg = my_sheet[llaveg1]
            llaveg.value = "FIRMA Y SELLO"
            llaveg.font = fuente2
            # g43 = my_sheet['G43']
            # g43.value = "FIRMA Y SELLO"
            # g43.font = fuente2
            llavek = 'k' + str(voy)
            llavek1 = 'K' + str(voy)
            llavek = my_sheet[llavek1]
            llavek.value = "FIRMA Y SELLO"
            llavek.font = fuente2
            # k43 = my_sheet['K43']
            # k43.value = "FIRMA Y SELLO"
            # k43.font = fuente2

            voy = voy + 1

            llaveb = 'b' + str(voy)
            llaveb1 = 'B' + str(voy)
            llaveb = my_sheet[llaveb1]
            llaveb.value = "NOTA ACLARATORIA. "
            llaveb.font = fuente1
            # b44 = my_sheet['B44']
            # b44.value = "NOTA ACLARATORIA. "
            # b44.font = fuente1

            llavee = 'e' + str(voy)
            llavee1 = 'E' + str(voy)
            llavee = my_sheet[llavee1]
            llavee.value = "TODA CANTIDAD RECIBIDA, MAYOR A LA SOLICITADA EN LA ORDEN DE COMPRA NO SERÁ PAGADA POR LA CLÍNICA MEDICAL S.A.S"
            llavee.font = fuente1

            # e44 = my_sheet['E44']
            # e44.value = "TODA CANTIDAD RECIBIDA, MAYOR A LA SOLICITADA EN LA ORDEN DE COMPRA NO SERÁ PAGADA POR LA CLÍNICA MEDICAL S.A.S"
            # e44.font = fuente2
            print("Pase Final con archivo =", archivo)

            response = HttpResponse(content_type="application/ms-excel")
            print("Pase Final1 con archivo =", archivo)
            contenido = "attachment; filename ={0}".format(archivo)
            print("Pase Final2 con archivo =", archivo)

            response["Content-Disposition"] = contenido
            print("Grabo archivo =", archivo)
            #print("Esta es la ruta actual=", os.getcwd())

            my_wb.save(archivo)

            # Fin rutina que crea el archivo excel

            # return render(self.request, self.template_name, context)
            print("Voy a cargar la pagina para mostrar el consecutivo delarchivo generado")
            return render(self.request, "Reportes/cabeza.html", context)

            ## Fin Extraemos los items de la Compra




            ## Fin rutina para reporte 5 Generar reportes Excel


        if (numeroReporte != 5):
            desdeFecha =  request.POST["desdeFechaSolicitud"]
            hastaFecha =  request.POST["hastaFechaSolicitud"]
            coordinador = request.POST["coordinador"]
            print ("desdeFecha = ",desdeFecha)
            print("hastafecha = ", hastaFecha)

        # Consigo Numero de Parametros del reporte
        encabezados=""
        cuerpo_sql = ""
        hayParametros= 0
        parametros = []

        if (numeroReporte=='1'):
            print ("Entre numeroReporte = 1")
            nombreReporte = "Solicitado Coordinador"
            hayParametros = 3
            cuerpo_sql = 'SELECT sol.id solicitud, sol.fecha fecha, area.area area,  usu.nom_usuario usuarioSolicitud, sol0.item item , des.nombre descripcion, pre.nombre presentacion, tipo.nombre tipoCompra, sol0.producto producto ,art.articulo articulo,sol0.cantidad cantidad, est.nombre estado, usucomp.nom_usuario usuarioCompra, "ordenCompra_id" ordenCompra from solicitud_solicitudes sol  left join solicitud_solicitudesdetalle sol0 on (sol0.solicitud_id = sol.id) inner join solicitud_usuarios usu on (usu.id= sol.usuarios_id) inner join solicitud_areas area on (area.id = sol.area_id) inner join solicitud_descripcioncompra des on (des.id = sol0.descripcion_id) inner join solicitud_presentacion pre on (pre.id = sol0.presentacion_id) inner join solicitud_tiposcompra tipo on (tipo.id = sol0."tiposCompra_id") inner join mae_articulos art on (art.codreg_articulo = sol0.producto) inner join solicitud_usuarios usucomp on (usucomp.id= sol0."usuarioResponsableCompra_id") inner join solicitud_estadosvalidacion est on (est.id = sol0."estadosCompras_id") WHERE sol.usuarios_id = ? and fecha >= ? and fecha <= ? ORDER BY fecha'
            encabezados = "#, fecha, area,usuSolicitud, item, desc,present, tipoCompra, producto , articulo, cantidad, estado, usuCompra,  ordenNo"
            parametros.append(coordinador)
            parametros.append(desdeFecha)
            parametros.append(hastaFecha)

            total = len(parametros)
            print("numero de parametros =", total)
            t = range(1, total + 1)
            for i in t:
                print("Matriz parametros = ", parametros[i - 1])
                dato = "'" + parametros[i - 1] + "'"
                cuerpo_sql = cuerpo_sql.replace("?", dato, 1)

        if (numeroReporte == '2'):
            print("Entre numeroReporte = 3")
            nombreReporte = "Validacion Ambito"
            hayParametros = 3
            cuerpo_sql = 'SELECT sol0.id solicitudNo,to_char(sol0.fecha,' + "'YYYY - MM - DD HH: MM.SS'" + ') fecha,  areas.area area,  usuariosCreaSol.nom_usuario usuariosCreaSol, sol.item item, des.nombre descripcion, tip.nombre tipo, sol.producto producto, art.articulo producto,pres.nombre  presentacion, sol.cantidad cantidad, sol.justificacion justificacion, sol."especificacionesTecnicas" tec, usu.nom_usuario usuResp,  est.nombre estValidacion FROM public.solicitud_solicitudes sol0 inner join  public.solicitud_solicitudesDetalle sol on (sol.solicitud_id=sol0.id) inner join public.solicitud_descripcioncompra des on (des.id = sol.descripcion_id ) inner join public.solicitud_tiposcompra tip on (tip.id = sol."tiposCompra_id" ) inner join public.solicitud_presentacion pres on (pres.id = sol.presentacion_id ) inner join public.mae_articulos art on (art.codreg_articulo = sol.producto) left join public.solicitud_usuarios usu on (usu.id = sol."usuarioResponsableValidacion_id") inner join public.solicitud_estadosvalidacion est on (est.id = sol."estadosValidacion_id") inner join public.solicitud_areas areas on (areas.id = sol0.area_id) inner join public.solicitud_usuarios usuariosCreaSol on (usuariosCreaSol.id = sol0.usuarios_id)  WHERE  sol."usuarioResponsableValidacion_id" = ? and sol0.fecha>= ? and  sol0.fecha<= ? ORDER BY sol0.fecha,sol0.id,sol.item '
            encabezados = "#, fecha, area, usuariosCreaSol, item, descripcion, tipo,  producto, art.articulo producto, presentacion,  cantidad, justificacion, tec, usuResp, estValidacion"
            parametros.append(coordinador)
            parametros.append(desdeFecha)
            parametros.append(hastaFecha)

            total = len(parametros)
            print("numero de parametros =", total)
            t = range(1, total + 1)
            for i in t:
                print("Matriz parametros = ", parametros[i - 1])
                dato = "'" + parametros[i - 1] + "'"
                cuerpo_sql = cuerpo_sql.replace("?", dato, 1)


        if (numeroReporte == '3'):

            print("Entre numeroReporte = 3")
            nombreReporte = "Almacen Coordinador"
            hayParametros = 3
            cuerpo_sql = 'SELECT sol0.id solicitudNo,to_char(sol0.fecha, ' + "'YYYY - MM - DD HH: MM.SS'" + ') fecha,  areas.area area,  usuariosCreaSol.nom_usuario usuariosCreaSol, sol.item item, des.nombre descripcion, tip.nombre tipo ,sol.producto producto,  art.articulo nombre_producto ,pres.nombre presentacion,sol.cantidad cantidad , sol.justificacion justificacion , sol."especificacionesTecnicas" tec, est.nombre estValidacion, est1.nombre estadosAlmacen,  sol."especificacionesAlmacen" especificacionesAlmacen,   usu1.nom_usuario usuAlmacen FROM public.solicitud_solicitudes sol0 INNER JOIN public.solicitud_solicitudesDetalle sol on (sol.solicitud_id=sol0.id) INNER JOIN public.solicitud_descripcioncompra des ON (des.id = sol.descripcion_id ) INNER JOIN public.solicitud_tiposcompra tip ON (tip.id = sol."tiposCompra_id" ) INNER JOIN public.solicitud_presentacion pres on (pres.id = sol.presentacion_id) INNER JOIN public.mae_articulos art   ON (art.codreg_articulo = sol.producto) LEFT JOIN  public.solicitud_usuarios usu ON (usu.id = sol."usuarioResponsableValidacion_id") LEFT JOIN public.solicitud_usuarios usu1 ON (usu1.id = sol."usuarioResponsableAlmacen_id") INNER JOIN public.solicitud_estadosvalidacion est ON (est.id = sol."estadosValidacion_id" ) INNER JOIN public.solicitud_estadosalmacen est1  ON (est1.id = sol."estadosAlmacen_id") inner join public.solicitud_areas areas on (areas.id = sol0.area_id) inner join public.solicitud_usuarios usuariosCreaSol on (usuariosCreaSol.id = sol0.usuarios_id)  WHERE  sol."usuarioResponsableAlmacen_id" =  ? and sol0.fecha >= ? and sol0.fecha <= ? ORDER BY sol0.fecha,sol0.id,sol.item '
            encabezados = "#, fecha,  area,usuariosCreaSol,  item, descripcion, tipo , producto, nombre_producto , presentacion, cantidad, justificacion  , tec, estValidacion, estadosAlmacen, especificacionesAlmacen,   usuAlmacen "
            parametros.append(coordinador)
            parametros.append(desdeFecha)
            parametros.append(hastaFecha)

            total = len(parametros)
            print("numero de parametros =", total)
            t = range(1, total + 1)
            for i in t:
                print("Matriz parametros = ", parametros[i - 1])
                dato = "'" + parametros[i - 1] + "'"
                cuerpo_sql = cuerpo_sql.replace("?", dato, 1)


        if (numeroReporte == '4'):

            print("Entre numeroReporte = 4")
            nombreReporte = "Compras Coordinador"
            hayParametros = 3
            cuerpo_sql = 'select ord.id orden, substring(to_char(ord."fechaElab",' + "'yyyy-mm-dd'" + '),1,10) fechaElab,area.area area, usu.nom_usuario usuarioCompras, proveedor proveedor,  sol.item item,art.articulo articulo, pre.nombre presenta, sol.iva iva,sol."solicitadoOrdenCantidad" solicitadoOrdenCantidad ,sol."recibidoOrdenCantidad" recibidoOrdenCantidad,sol."valorUnitario" valorUnitario,sol."solicitadoOrdenValor" solicitadoOrdenValor,sol."recibidoOrdenValor" recibidoOrdenValor ,ord."valorBruto" valorBruto,ord."descuento" descuento,ord."valorParcial" valorParcial, ord."iva" iva, ord."valorTotal" valorTotal,case when ord."estadoOrden" = ' + "'V'" + ' then ' + "'Vigente'" + ' when ord."estadoOrden" = ' + "'C'" + ' then ' +  "'caduca'" + ' end  estadoOrden ,case when ord.opciones= ' + "'A'" + ' then ' + "'Anticipo'" + ' when ord.opciones= ' + "'N'" + ' then ' + "'Noventa dias'" + ' when ord.opciones= ' + "'C'" + ' then ' + "'Contra enrega'" + '  end   opciones,ord.observaciones observaciones, usu1.nom_usuario usuarioAproboStaff FROM solicitud_ordenesCompra ord INNER JOIN solicitud_solicitudesdetalle sol ON (sol."ordenCompra_id" = ord.id) INNER JOIN solicitud_articulos art ON ( art."codregArticulo" = sol.producto) INNER JOIN solicitud_proveedores prov on (prov.id = ord.proveedor_id) INNER JOIN solicitud_areas area on (area.id = ord.area_id) INNER JOIN solicitud_usuarios usu on (usu.id = ord."responsableCompra_id") INNER JOIN solicitud_Staff usu1 on (usu1.id = ord."aproboCompraStaff_id") INNER JOIN solicitud_descripcioncompra des on (des.id = sol.descripcion_id) INNER JOIN solicitud_presentacion pre on (pre.id = sol.presentacion_id) WHERE ord."responsableCompra_id" = ? and ord."fechaElab"  >= ? and ord."fechaElab"  <= ? ORDER BY ord."fechaElab", sol.item'
            encabezados = "#,Elab,area, usuCompras,proveedor, item, articulo, presenta, iva, solicitadoOrdenCantidad , recibidoOrdenCantidad, valorUnitario, solicitadoOrdenValor, recibidoOrdenValor ,valorBruto,descuento, valorParcial, iva, valorTotal,  estadoOrden , opciones,observaciones,  usuarioAproboStaff"
            parametros.append(coordinador)
            parametros.append(desdeFecha)
            parametros.append(hastaFecha)

            total = len(parametros)
            print("numero de parametros =", total)
            t = range(1, total + 1)
            for i in t:
                print("Matriz parametros = ", parametros[i - 1])
                dato = "'" + parametros[i - 1] + "'"
                cuerpo_sql = cuerpo_sql.replace("?", dato, 1)


        print("CuerpoSQl_FINAL = ", cuerpo_sql)
        print("cuerpo_sql = ", cuerpo_sql)
        print("encabezados = ", encabezados)
        print("hayParametros = ", hayParametros)
        print("numeroReporte = ", numeroReporte)
        print("tipoArchivo = ", tipoArchivo)
        print("username = ", username)
        print("sedeSeleccionada =", sedeSeleccionada)

        print("numeroReporte =", numeroReporte)

        context['username'] = username
        context['sedeSeleccionada'] = sedeSeleccionada
        context['NombreSede'] = nombreSede
        context['NombreUsuario'] = nombreUsuario
        context['Perfil'] = perfil

        print ("hayParametros =", hayParametros)
        # Selecciono o escojo el cuerpo_sql

        print("CuerpoSQl_Final = ", cuerpo_sql)

        #desdeFecha = request.POST.get('DesdeFecha', False)
        #hastaFecha = request.POST.get('HastaFecha', False)

        # Aqui hace la pregunta si es excel o pdf



        # Story = [Spacer(0, 20)]
        buff = io.BytesIO()
        # doc = SimpleDocTemplate(buff, pagesize=letter, rightMargin=26,   leftMargin=32, topMargin=72, bottomMargin=18)
        doc = SimpleDocTemplate(buff, pagesize=letter, rightMargin=26, leftMargin=32, topMargin=120, bottomMargin=20)

        styles = getSampleStyleSheet()
        styleBH = styles["Normal"]
        styleBH.align = 'CENTER'
        styleBH.fontSize = 6

        estilos = getSampleStyleSheet()
        headline_mayor = estilos["Heading1"]
        headline_mayor.alignment = TA_LEFT
        headline_mayor.leading = 8
        headline_mayor.fontSize = 10
        headline_mayor.fontName = "Helvetica-Bold"
        headline_mayor.spaceAfter = 0
        headline_mayor.spaceBefore = 0

        headline_mayor1 = estilos["Heading5"]
        headline_mayor1.alignment = TA_LEFT
        headline_mayor1.leading = 6
        headline_mayor1.fontSize = 8
        headline_mayor1.fontName = "Helvetica-Bold"
        headline_mayor1.spaceAfter = 0
        headline_mayor1.spaceBefore = 0

        headline_mayor2 = estilos["Heading5"]
        headline_mayor2.alignment = TA_LEFT
        headline_mayor2.leading = 7
        headline_mayor2.fontSize = 8
        headline_mayor2.fontName = "Helvetica-Bold"
        headline_mayor2.spaceAfter = 0
        headline_mayor2.spaceBefore = 0

        headline_mayor3 = estilos["Heading5"]
        headline_mayor3.alignment = TA_CENTER
        headline_mayor3.leading = 8
        headline_mayor3.fontSize = 10
        headline_mayor3.fontName = "Helvetica-Bold"
        headline_mayor3.spaceAfter = 0
        headline_mayor3.spaceBefore = 0

        headline_mayor33 = estilos["Heading5"]
        headline_mayor33.alignment = TA_CENTER
        headline_mayor33.leading = 3
        headline_mayor33.fontSize = 10
        headline_mayor33.fontName = "Helvetica-Bold"
        headline_mayor33.spaceAfter = 0
        headline_mayor33.spaceBefore = 0

        headline_mayor4 = estilos["Heading5"]
        headline_mayor4.alignment = TA_CENTER
        # headline_mayor4.leftIndent= 10
        headline_mayor4.leading = 7
        headline_mayor4.fontSize = 9
        headline_mayor4.fontName = "Helvetica-Bold"
        headline_mayor4.spaceAfter = 0
        headline_mayor4.spaceBefore = 0

        subtitle_tipoevol = estilos["Heading2"]
        subtitle_tipoevol.leading = 9  # estaba15
        subtitle_tipoevol.fontSize = 8
        subtitle_tipoevol.fontName = "Times-Roman"
        subtitle_tipoevol.spaceAfter = 0
        subtitle_tipoevol.spaceBefore = 0
        subtitle_tipoevol.alignment = TA_LEFT
        subtitle_tipoevol.wordWrap = 'LTR'

        # subtitle_atencion = estilos["Heading3"]
        # subtitle_atencion.leading =9
        # subtitle_atencion.fontSize = 8
        # subtitle_atencion.fontName = "Times-Roman"
        # subtitle_atencion.spaceAfter = 0
        # subtitle_atencion.spaceBefore = 0
        # subtitle_atencion.alignment = TA_LEFT

        subtitle_atencion = estilos["Heading3"]
        subtitle_atencion.leading = 9
        subtitle_atencion.fontSize = 8
        subtitle_atencion.fontName = "courier-bold"
        subtitle_atencion.spaceAfter = 0
        subtitle_atencion.spaceBefore = 0
        subtitle_atencion.alignment = TA_LEFT
        # Tahoma ,, courier

        subtitle_cabezote = estilos["Heading4"]
        subtitle_cabezote.leading = 7
        subtitle_cabezote.fontSize = 8
        subtitle_cabezote.fontName = "Times-Roman"
        subtitle_cabezote.spaceAfter = 0
        subtitle_cabezote.spaceBefore = 0
        subtitle_cabezote.alignment = TA_LEFT

        subtitle_nacimiento = estilos["Heading6"]
        subtitle_nacimiento.leading = 7
        subtitle_nacimiento.fontSize = 8
        subtitle_nacimiento.fontName = "Times-Roman"
        subtitle_nacimiento.spaceAfter = 0
        subtitle_nacimiento.spaceBefore = 0
        subtitle_nacimiento.alignment = TA_LEFT

        estilos.add(ParagraphStyle(name='Justify', alignment=TA_RIGHT))
        estilos1 = getSampleStyleSheet()
        estilos1.add(ParagraphStyle(name='Justify_left', alignment=TA_LEFT))
        estilos2 = getSampleStyleSheet()
        estilos2.add(ParagraphStyle(name='Justify_right', alignment=TA_RIGHT))
        response    = HttpResponse(content_type='application/pdf')
        print("Creo Archivo")


        #response = HttpResponse(content_type="application/pdf")

        nombreReporteFinal = nombreReporte + ".pdf"
        response['Content-Disposition'] = 'attachment; filename= '  + nombreReporteFinal

        ## Controlo si genero o no el listado del informe

        print("A CONTROLAR EL ACCESO AL REPORTE")
        data = "usuario no tiene permisos para generar el informe"

        if (numeroReporte == '2'):
            if (perfil != 'V'):
                messages.warning(request, 'usuario no tiene permisos para generar el informe.')
                # return redirect('post_storeReportesConsulta',kwargs={'username': username, 'sedeSeleccionada': sedeSeleccionada,'nombreUsuario': nombreUsuario, 'nombreSede': nombreSede, 'perfil': perfil, 'numeroReporte':4})
                # return JsonResponse("usuario no tiene permisos para generar el informe")
                return HttpResponse(json.dumps(data), content_type="application/json")

        if (numeroReporte == '3'):
            if (perfil != 'A'):
                messages.warning(request, 'usuario no tiene permisos para generar el informe.')
                #return redirect('post_storeReportesConsulta',kwargs={'username': username, 'sedeSeleccionada': sedeSeleccionada,'nombreUsuario': nombreUsuario, 'nombreSede': nombreSede, 'perfil': perfil, 'numeroReporte':4})
                #return JsonResponse("usuario no tiene permisos para generar el informe")
                return HttpResponse(json.dumps(data), content_type="application/json")

        if (numeroReporte == '4'):
            if (perfil != 'C'):
                print("pase1")
                messages.warning(request, 'usuario no tiene permisos para generar el informe.')
                print("pase2")
                #return redirect('post_storeReportesConsulta',kwargs={'username': username, 'sedeSeleccionada': sedeSeleccionada ,'nombreUsuario': nombreUsuario, 'nombreSede': nombreSede, 'perfil': perfil, 'numeroReporte':4 })

                #return JsonResponse("usuario no tiene permisos para generar el informe")
                return HttpResponse(json.dumps(data), content_type="application/json")


        ## FIN CONROLO ACCESO

        #response['Content-Disposition'] = 'attachment; filename="' + tipodoc + ' ' + documento + '.pdf"'

        ## Aqui va la impresion de todo el reporte
        ##
        ##


        ## Genero el Reporte Dinamico

        miConexion = psycopg2.connect(host="192.168.0.237", database="bd_solicitudes0", port="5432", user="postgres",
                                      password="BD_m3d1c4l")
        miConexion.set_client_encoding('LATIN1')

        cur = miConexion.cursor()

        cur.execute("set client_encoding='LATIN1';")
        print(cur.execute("show server_encoding;"))

        # Aqui se arregla Cuerpo_sql, con los parametros y los valores introducidos por el Usuario


        cur.execute(cuerpo_sql)
        print("Esto lo ejecuto = ", cuerpo_sql)
        rows = cur.fetchall()

        #Story.append(Spacer(1, 3))
        #print ("Este es el listado de Registros " , rows)

        # ExtraigoCuento cuantas columnas hay
        global columnas
        columnas = encabezados.count(',')
        columnas = columnas + 1

        print ("El numero de columnas del reporte son :",  columnas)

        # Extraigo el valor de los encabezados

        t = ","
        #encabezado = "codreg_sede, nom_sede, codreg_ips, direccion, telefono, departamento"
        #columnas = 6

        posicioncoma=0
        initial = encabezados
        global lasColumnas
        lasColumnas = []


        for i in range(columnas):

            posicioncoma =  initial.find(t)
            #print("Posicion coma = ", posicioncoma)
            lasColumnas.append(initial[0:(posicioncoma)])
            initial = initial[(posicioncoma +1) : len(encabezados)]
            #print("initial = ", initial)

        print("Estas son las Columnas :",    lasColumnas)



        if (tipoArchivo == "csv"):

            response = HttpResponse(content_type='text/csv')

            print ("nombreReporte = ", nombreReporte)

            nombreReporteFinal = nombreReporte + ".csv"
            response['Content-Disposition'] = 'attachment; filename= ' + nombreReporteFinal

            print("nombreReporteFinal = ", nombreReporteFinal)


            myFile = open(nombreReporteFinal, 'w')

            with myFile:
                writer = aliascsv.writer(response, myFile)

            response.write(u'\ufeff'.encode('utf8'))


            # write column headers in sheet
            titulos = ""

            for col_num in range(len(lasColumnas)):
                titulos = titulos +   lasColumnas[col_num] + ","

            writer.writerow([
                smart_str(titulos),
            ])

            row_num = 0

            if rows == []:
                pass
            else:

                for row in rows:
                    row_num = row_num + 1
                    campo= ""
                    campoTot= ""

                    for col in range(len(lasColumnas)):
                        campo = row[col]

                        campoTot = campoTot  + str(campo) + ","

                    writer.writerow([        smart_str(campoTot),     ])

        if (tipoArchivo == "grilla"):

            grilla_data = []

            subir= {}
            x = range(0, len(lasColumnas))

            for row in rows:
               subir = {}
               for j in x:

                   subir[lasColumnas[j].lstrip()] = str(row[j])

               grilla_data.append(subir)

            #print (grilla_data)

            sonColumnas = range(0, len(lasColumnas))

            context['Grilla'] = rows # grilla_data
            context['LasColumnas'] = lasColumnas
            context['NumeroColumnas'] = sonColumnas
            context['Username'] = username
            context['SedeSeleccionada'] = sedeSeleccionada


            context['NombreSede'] = nombreSede

            return render(request, "Reportes/PantallaGrilla.html", context)

        if (tipoArchivo == "grafica"):

            print("Entre GRAFICA");
            username = request.POST["username"]
            context['username'] = username
            sedeSeleccionada = request.POST["sedeSeleccionada"]
            context['sedeSeleccionada'] = sedeSeleccionada

            colGraf = request.POST["campos"]
            print ("colGraf = ", colGraf)
            colTipo = request.POST["campos1"]
            print("colTipo = ", colTipo)
            row_num = 0

            # load data into a DataFrame object:

            valores = []

            if rows == []:
                pass
            else:
                for row in rows:
                    row_num = row_num + 1
                    campo = ""
                    campoTot = ""

                    for col in range(len(lasColumnas)):
                       if (lasColumnas[col].strip() == colGraf):
                           busco="datetime"
                           hayfec= lasColumnas[col].find(busco)
                           if (hayfec==0):
                                s = re.sub(r"([^n\u0300-\u036f]|n(?!\u0303(?![\u0300-\u036f])))[\u0300-\u036f]+", r"\1",
                                    normalize("NFD", row[col]), 0, re.I)
                                s = normalize('NFC', s)
                           else:
                               s=row[col]

                           #valores.append({"items": row[col], colGraf: row[col]})
                           valores.append({"items": row[col], colGraf: s})
                           print ("Entre columna", lasColumnas[col])

                print ("valores", valores)

            df = pd.DataFrame(valores, columns=['items', colGraf])
            print(df)
            bins = [0, 5, 12, 18, 35, 60, 100]
            names = ["0-5", "6-12", "13-18", "19-35", "36-60", "mas de 60"]
            #df['edad'] = pd.cut(df['edad'], bins, labels = names)

            if df.empty:
                return HttpResponse ("No hay Infomacion")

            datos = df.groupby([colGraf], as_index=False).agg({'items': 'count'})

            print("datos: ", datos)
            if colTipo=="Barras":
                print("Entre Barras")
                plt.bar( datos[colGraf],   datos['items'])
                plt.grid(color='#95a5a6', linestyle='--', linewidth=2, axis='y', alpha=0.7)
                plt.title(nombreReporte)
                plt.legend(datos,loc=1)
                plt.xlabel(colGraf)
                plt.ylabel("Ocurrencias")
            else:
                print ("Entre Pie")
                plt.pie(datos['items'], labels=datos[colGraf], autopct='%.0f%%', shadow = True)
                plt.title(nombreReporte)
                plt.legend()
                plt.legend(title=colGraf, loc ="center right")

            print("Antes de Reporte = ")
            reporte = "{% static '/img/myfig.jpeg' %}"
            print("Reporte = ", reporte)
            plt.savefig("C:\EntornosPython\MedicalReportes8\MedicalReportes8\static\img\myfig.jpeg", format = 'jpeg')
            #plt.savefig("myfig.jpg", format='jpeg')
            print("ya grave imagen")
            #plt.show(block=True)
            plt.interactive(False)

            # Consigo Nombre de la sede

            miConexion = psycopg2.connect(host="192.168.0.237", database="bd_solicitudes0", port="5432", user="postgres",
                                          password="BD_m3d1c4l")
            cur = miConexion.cursor()
            comando = "SELECT codreg_sede, nom_sede FROM imhotep_sedes WHERE codreg_sede = '" + sedeSeleccionada + "'"
            cur.execute(comando)
            print(comando)

            nombreSede = []

            for codreg_sede, nom_sede in cur.fetchall():
                nombreSede.append({'codreg_sede': codreg_sede, 'nom_sede': nom_sede})

            miConexion.close()

            context['NombreSede'] = nombreSede

            # Le doy la informacion de los reportes a los que tiene acceso

            miConexion = psycopg2.connect(host="192.168.0.237", database="bd_solicitudes0", port="5432", user="postgres",
                                          password="BD_m3d1c4l")

            miConexion.set_client_encoding('LATIN1')
            cur = miConexion.cursor()
            cur.execute("set client_encoding='LATIN1';")
            # comando = 'select  reportes.id numreporte, usuarios.cod_usuario usuario, reportes.nom_reporte reporte,reportes.cuerpo_sql, reportes.descripcion descripcion , reportes.encabezados encabezados from public."Administracion_mae_repusuarios" as usuarios,  public."Administracion_mae_reportes" as reportes , public."Administracion_imhotep_sedesreportes" sedes  where usuarios.cod_Usuario = ' + "'" + username + "'" + ' and  usuarios.mae_reportes_id = reportes.id  and usuarios.cod_sede_id = sedes.id and sedes.codreg_sede = '  + "'"  + sedeSeleccionada + "'"
            # comando = 'select  reportes.id numreporte, usuarios.cod_usuario usuario, reportes.nom_reporte reporte,reportes.cuerpo_sql, reportes.descripcion descripcion , reportes.encabezados encabezados from public."Administracion_mae_repusuarios" as usuarios,  public."Administracion_mae_reportes" as reportes , public."Administracion_imhotep_sedesreportes" sedes  where usuarios.cod_Usuario = ' + "'" + username + "'" + ' and  usuarios.mae_reportes_id = reportes.id  and usuarios.cod_sede_id = sedes.id and sedes.codreg_sede = ' + "'" + sedeSeleccionada + "'" + ' AND usuarios.estadoReg=' +  "'A'"
            comando = 'select  reportes.id numreporte, usuarios.cod_usuario usuario, reportes.nom_reporte reporte,reportes.cuerpo_sql, reportes.descripcion descripcion , reportes.encabezados encabezados ,reportes.mae_gruporeportes_id grupo ,reportes.mae_subgruporeportes_id subgrupo , grupos.nom_grupo nombreGrupo, subgrupos.nom_subgrupo nombreSubgrupo from public."Administracion_mae_repusuarios" as usuarios,  public."Administracion_mae_reportes" as reportes , public."Administracion_imhotep_sedesreportes" sedes ,public."Administracion_mae_gruporeportes" grupos,public."Administracion_mae_subgruporeportes" subgrupos   where usuarios.cod_Usuario = ' + "'" + username + "'" + ' and  usuarios.mae_reportes_id = reportes.id  and usuarios.cod_sede_id = sedes.id and grupos.id = reportes.mae_gruporeportes_id and grupos.id = ' + "'" + grupo + "'" + ' and subgrupos.id = reportes.mae_subgruporeportes_id  AND subgrupos.id = ' + "'" + subGrupo + "'" + ' and sedes.codreg_sede = ' + "'" + sedeSeleccionada + "'" + ' AND usuarios.estadoReg=' + "'A'" + ' AND reportes.estadoReg=' + "'A'"

            print(comando)
            cur.execute(comando)

            reportesUsuario = []

            for numreporte, usuario, reporte, cuerpo_sql, descripcion, encabezados, grupo, subgrupo, nombreGrupo, nombreSubGrupo in cur.fetchall():
                reportesUsuario.append(
                    {'numreporte': numreporte, 'usuario': usuario, 'reporte': reporte, 'cuerpo_sql': cuerpo_sql,
                     'descripcion': descripcion, 'encabezados': encabezados, 'grupo': grupo, 'subgrupo': subgrupo,
                     'nombreGrupo': nombreGrupo, 'nombreSubGrupo': nombreSubGrupo})

            miConexion.close()
            context['ReportesUsuario'] = reportesUsuario

            plt.close('all')


            response = HttpResponse(content_type="application/ms-excel")

            img = open('C:\EntornosPython\MedicalReportes8\MedicalReportes8\static\img\myfig.jpeg', 'rb')

            response = FileResponse(img)

            return response


        if (tipoArchivo == "excel"):

            response = HttpResponse(content_type="application/ms-excel")
            nombreReporteFinal = nombreReporte + ".xls"
            response['Content-Disposition'] = 'attachment; filename=' + nombreReporteFinal

            # creating workbook
            wb = xlwt.Workbook(encoding='utf-8')

            # adding sheet
            #Info0 = wb.add_sheet("Info0")

            styles = dict(
                bold='font: bold 1',
                italic='font: italic 1',
                # Wrap text in the cell
                wrap_bold='font: bold 1; align: wrap 1;',
                # White text on a blue background
                reversed='pattern: pattern solid, fore_color blue; font: color white;',
                # Light orange checkered background
                light_orange_bg='pattern: pattern fine_dots, fore_color white, back_color orange;',
                # Heavy borders
                bordered='border: top thick, right thick, bottom thick, left thick;',
                # 16 pt red text
                big_red='font: height 260, color blue;',

                # 16 pt red text
                normal='font: height 260, color black;',
            )

           # for idx, k in enumerate(sorted(styles)):
           #     style = xlwt.easyxf(styles[k])
           #     ws.write(idx, 0, k)
           #     ws.write(idx, 1, styles[k], style)


            # Sheet header, first row
            row_num = 0

            # Encabezados del Reporte

            #font_style = xlwt.XFStyle()
            # headers are bold
            #font_style.font.bold = True

            #row_num = 1

            #Info0.write(row_num, 0, "CLINICA MEDICAL", font_style)
            #row_num = row_num + 1
            #Info0.write(row_num, 0, 'NIT: 8305077188', font_style)
            #row_num = row_num + 1
            #Info0.write(row_num, 0, "INFORME: ", font_style)
            #Info0.write(row_num, 1, nombreReporte, font_style)
            #Info0.write(row_num, 3, "FECHA: ", font_style)
            #fechaActual = datetime.today().strftime('%Y-%m-%d %H:%M')
            #print("Fecha Actual = ", fechaActual)
            #Info0.write(row_num, 4, fechaActual, font_style)

            #row_num = row_num + 2

            # write column headers in sheet
            #for col_num in range(len(lasColumnas)):
            #    style = xlwt.easyxf(styles['big_red'])
            #    #ws.write(idx, 1, styles[k], style)
            #    Info0.write(row_num, col_num, lasColumnas[col_num], style)

            #row_num = row_num + 1

            # Sheet body, remaining rows
            #font_style = xlwt.XFStyle()
            #font_style.font.bold = True

            # get your data, from database or from a text file...
            global Info0
            if rows == []:
                A1=0
                Info0 = wb.add_sheet("Info0")
            else:
                style = xlwt.easyxf(styles['normal'])

                # Practicamente desde aquio se comienza a imprimir el reporte

                print("ESte es el tamaño de que ? " , len(rows))

                numeroHojas = math.trunc(len(rows)/60000)

                if numeroHojas == 0:
                    numeroHojas = 1
                else:
                     if (numeroHojas % numeroHojas != 1):
                          numeroHojas = numeroHojas + 1



                print("El Numero de Hojas =  ", math.trunc(numeroHojas))

                x = range(0, (numeroHojas))

                nombreDeHojas = []
                Info = "Info"

                resultado = 0

                n = range(1, numeroHojas + 1)

                for hoj in n:
                    if (hoj == 1):

                        print("Entre a crear la Primera Info0")

                        Info0 = wb.add_sheet("Info0")
                    if (hoj == 2):
                        global Info1
                        Info1 = wb.add_sheet("Info1")
                    if (hoj == 3):
                        global Info2
                        Info2 = wb.add_sheet("Info2")
                    if (hoj == 4):
                        global Info3
                        Info3 = wb.add_sheet("Info3")
                    if (hoj == 5):
                        global Info4
                        Info4 = wb.add_sheet("Info4")
                    if (hoj == 6):
                        global Info5
                        Info5 = wb.add_sheet("Info5")
                    if (hoj == 7):
                        global Info6
                        Info6 = wb.add_sheet("Info6")
                    if (hoj == 8):
                        global Info7
                        Info7 = wb.add_sheet("Info7")
                    if (hoj == 9):
                        global Info8
                        Info8 = wb.add_sheet("Info8")
                    if (hoj == 10):
                        global Info9
                        Info9 = wb.add_sheet("Info9")
                    if (hoj == 11):
                        global Info10
                        Info10 = wb.add_sheet("Info10")
                    if (hoj == 12):
                        global Info11
                        Info11 = wb.add_sheet("Info11")
                    if (hoj == 13):
                        global Info12
                        Info12 = wb.add_sheet("Info12")
                    if (hoj == 14):
                        global Info13
                        Info13 = wb.add_sheet("Info13")
                    if (hoj == 15):
                        global Info14
                        Info14 = wb.add_sheet("Info14")
                    if (hoj == 16):
                        global Info15
                        Info15 = wb.add_sheet("Info15")
                    if (hoj == 17):
                        global Info16
                        Info16 = wb.add_sheet("Info16")
                    if (hoj == 18):
                        global Info17
                        Info17 = wb.add_sheet("Info17")
                    if (hoj == 19):
                        global Info18
                        Info18 = wb.add_sheet("Info18")
                    if (hoj == 20):
                        global Info19
                        Info19 = wb.add_sheet("Info19")

                if numeroHojas <= 1:
                        desde=0
                        hasta=len(rows)

                if numeroHojas > 1:
                        desde = 0
                        hasta = 60000
                        # Aqui debe crear las hojas que va a Utilizar

                for z in x:
                    # Aqui impresion de titulos,

                    if (z==0):

                        resultado = titulosCab(Info0)
                    if (z==1):
                        resultado = titulosCab(Info1)
                    if (z==2):
                        titulosCab(Info2)
                    if (z==3):
                        titulosCab(Info3)
                    if (z==4):
                        titulosCab(Info4)
                    if (z==5):
                        titulosCab(Info5)
                    if (z==6):
                        titulosCab(Info6)
                    if (z==7):
                        titulosCab(Info7)
                    if (z==8):
                        titulosCab(Info8)
                    if (z==10):
                        titulosCab(Info10)
                    if (z==11):
                        titulosCab(Info11)
                    if (z==12):
                        titulosCab(Info12)
                    if (z==13):
                        titulosCab(Info13)
                    if (z==14):
                        titulosCab(Info14)
                    if (z==15):
                        titulosCab(Info15)
                    if (z==16):
                        titulosCab(Info16)
                    if (z==17):
                        titulosCab(Info17)
                    if (z==18):
                        titulosCab(Info18)
                    if (z==19):
                        titulosCab(Info19)


                    for i in range(desde, hasta):
                        print("Ya casi 00")
                        for j in range(0, len(rows[i])):
                            print("Ya casi 01")
                            if z==0:
                                print("Ya casi 02")
                                print("fila = ", i+8)
                                print("columna = ", j)
                                print("valor = ", rows[i][j])
                                Info0.write(i+8, j, str(rows[i][j]), style)

                            if z==1:
                                Info1.write(i-desde+8, j, str(rows[i][j]), style)
                            if z==2:
                                Info2.write(i-desde+8, j, str(rows[i][j]), style)
                            if z==3:
                                Info3.write(i-desde+8, j, str(rows[i][j]), style)
                            if z==4:
                                Info4.write(i-desde+8, j, str(rows[i][j]), style)
                            if z==5:
                                Info5.write(i-desde+8, j, str(rows[i][j]), style)
                            if z==6:
                                Info6.write(i-desde+8, j, str(rows[i][j]), style)
                            if z==7:
                                Info7.write(i-desde+8, j, str(rows[i][j]), style)
                            if z==8:
                                Info8.write(i-desde+8, j, str(rows[i][j]), style)
                            if z==9:
                                Info9.write(i-desde+8, j, str(rows[i][j]), style)
                            if z==10:
                                Info10.write(i-desde+8, j, str(rows[i][j]), style)
                            if z==11:
                                Info11.write(i-desde+8, j, str(rows[i][j]), style)
                            if z==12:
                                Info12.write(i-desde+8, j, str(rows[i][j]), style)
                            if z==13:
                                Info13.write(i-desde+8, j, str(rows[i][j]), style)
                            if z==14:
                                Info14.write(i-desde+8, j, str(rows[i][j]), style)
                            if z==15:
                                Info15.write(i-desde+8, j, str(rows[i][j]), style)
                            if z==16:
                                Info16.write(i-desde+8, j, str(rows[i][j]), style)
                            if z==17:
                                Info17.write(i-desde+8, j, str(rows[i][j]), style)
                            if z==18:
                                Info18.write(i-desde+8, j, str(rows[i][j]), style)
                            if z==19:
                                Info19.write(i-desde+8, j, str(rows[i][j]), style)


                    desde = hasta

                    print("Ya casi 1")
                    if (z == (numeroHojas-2)):
                        print("Ya casi 2")
                        hasta = len(rows)
                        print ("Hasta de Control = " , hasta)
                        print("z =  ", z)
                    else:
                        print("Ya casi 3")
                        hasta = hasta + 60000
                        print("Hasta Normal = ", hasta)
                        print("z =  ", z)


                #for row in rows:
                #    row_num = row_num + 1
                #    print ("fila : ", row)


                #    for col in range(len(lasColumnas)):

                #       campo = row[col]
                #        print("fila= ", row_num)
                #        print("columna= ", col)
                #        print ("campo= ", campo)
                #        ws.write(row_num, col, campo, style)



        if (tipoArchivo=='pdf'):

            # Genera el pdf


            if rows == []:

                tbl_data4 = ['  ']
                print ("Entre por No registros PDF")
                tbl1 = Table(tbl_data4, colWidths =[10 * cm, 1.6 * cm, 1.4  * cm, 1  * cm, 4.6  * cm, 1 * cm])

                Story.append(tbl1)
                Story.append(Spacer(1, 3))

            else:
                print("Entre por SI HAY  registros PDF")
                # Aqui Rutina de Impresion de Titulos

                comienzo = 0
                mascara = ""
                print ("Aqui va el listado de Registros")

                for row in rows:
                    tbl_data3 = []
                    tbl_data2 = []
                    tbl_data1 = []
                    longitudes = []
                    longitudesFinal = []
                    longitudesFinal1 = {}
                    longitudesFinal1['formato'] = ""
                    Ancho = 0
                    calculo = 0
                    son = ""
                    print("fila : ", row)
                    print ("columnas = ", columnas)

                    m = range(0, columnas)

                    for i in m:
                            print("la variable i = ", i)
                            print("columna longitud = ", len(str(row[i])))
                            longitudes.append(len(str(row[i])))
                            Ancho = Ancho + int(len(str(row[i])))

                    print ("Total Ancho Columnas =", Ancho)

                    for i in m:
                            tbl_data1 = Paragraph(str(row[i]), subtitle_tipoevol),
                            tbl_data2.append(tbl_data1)
                            calculo = round(longitudes[i] * 19 / Ancho,2)

                            if calculo < 1:
                                calculo = 1
                            son = son + str(calculo) + " * cm, "
                            #longitudesFinal1['formato'] = longitudesFinal1['formato'] + str(calculo) + " * cm, "

                    longitudesFinal.append(son.replace("'",''))
                    #longitudesFinal1['formato']  = son

                    son = '[' + son + ']'
                    son.replace("'", ' ')
                    print('son = ', son)
                    print("longitud Final = ", longitudesFinal)
                    print("longitud Final Otro = ", longitudesFinal[0])

                    print ("tbl_data2 = ", tbl_data2)
                    tbl_data3.append(tbl_data2)
                    print("tbl_data3 = ", tbl_data3)

                    #tbl1 = Table(tbl_data3, colWidths=[3.5 * cm, 2   * cm, 8    * cm, 2    * cm, 3    * cm, 0.5  * cm])
                    #floats = list(map(float, longitudesFinal1['formato']))


                    tbl1 = Table(tbl_data3, colWidths=None)
                    #tbl1 = Table(tbl_data3, colWidths=floats)

                    Story.append(tbl1)
                    Story.append(Spacer(1, 3))

        miConexion.close()


        # Le doy la informacion de los reportes a los que tiene acceso

        miConexion = psycopg2.connect(host="192.168.0.238", database="bd_imhotep", port="5432", user="postgres",
                                      password="BD_m3d1c4l")
        miConexion.set_client_encoding('LATIN1')
        cur = miConexion.cursor()
        cur.execute("set client_encoding='LATIN1';")
        comando = 'select  reportes.id numreporte, usuarios.cod_usuario usuario, reportes.nom_reporte reporte,reportes.cuerpo_sql, reportes.descripcion descripcion , reportes.encabezados encabezados from public."Administracion_mae_repusuarios" as usuarios,public."Administracion_mae_reportes" as reportes , public."Administracion_imhotep_sedesreportes" sedes  where usuarios.cod_Usuario = ' + "'" + username + "'" + ' and  usuarios.mae_reportes_id = reportes.id and sedes.codreg_sede = ltrim(' + "'" + str(sedeSeleccionada) + "')" + ' AND reportes.estadoReg=' + "'A'"

        print(comando)
        cur.execute(comando)

        reportesUsuario = []

        for numreporte, usuario, reporte, cuerpo_sql, descripcion, encabezados in cur.fetchall():
            reportesUsuario.append(
                {'numreporte': numreporte, 'usuario': usuario, 'reporte': reporte, 'cuerpo_sql': cuerpo_sql,
                 'descripcion': descripcion, 'encabezados': encabezados})

        miConexion.close()

        print("pase0")

        context['ReportesUsuario'] = reportesUsuario


        if (tipoArchivo == "excel"):
            print("vOY DE REGRESO CON EL eXCEL")
            wb.save(response)
            print("vOY DE REGRESO CON EL eXCEL1")
            return response

        if (tipoArchivo == "pdf"):
            print("Voy a generar el reporte")
            doc.build(Story, onFirstPage=self.myFirstPage, onLaterPages=self.myLaterPages)
            response.write(buff.getvalue())
            buff.close()

            return response

        if (tipoArchivo == "csv"):

            return response


def titulosCab(RecibeInfo):

        print("Entre Rutina de impresion")

        row_num = 0

        styles = dict(
            bold='font: bold 1',
            italic='font: italic 1',
            # Wrap text in the cell
            wrap_bold='font: bold 1; align: wrap 1;',
            # White text on a blue background
            reversed='pattern: pattern solid, fore_color blue; font: color white;',
            # Light orange checkered background
            light_orange_bg='pattern: pattern fine_dots, fore_color white, back_color orange;',
            # Heavy borders
            bordered='border: top thick, right thick, bottom thick, left thick;',
            # 16 pt red text
            big_red='font: height 260, color blue;',

            # 16 pt red text
            normal='font: height 260, color black;',
        )

        row_num = 1

        font_style = xlwt.XFStyle()
        # headers are bold
        font_style.font.bold = True

        RecibeInfo.write(row_num, 0, "CLINICA MEDICAL", font_style)
        row_num = row_num + 1
        RecibeInfo.write(row_num, 0, 'NIT: 8305077188', font_style)
        row_num = row_num + 1
        RecibeInfo.write(row_num, 0, "INFORME: ", font_style)
        RecibeInfo.write(row_num, 1, nombreReporte, font_style)
        RecibeInfo.write(row_num, 3, "FECHA: ", font_style)
        fechaActual = datetime.today().strftime('%Y-%m-%d %H:%M')
        print("Fecha Actual = ", fechaActual)
        RecibeInfo.write(row_num, 4, fechaActual, font_style)

        row_num = row_num + 2

        # write column headers in sheet
        for col_num in range(len(lasColumnas)):
            style = xlwt.easyxf(styles['big_red'])
            # ws.write(idx, 1, styles[k], style)
            RecibeInfo.write(row_num, col_num, lasColumnas[col_num], style)
        row_num = row_num + 1
        # Sheet body, remaining rows
        font_style = xlwt.XFStyle()
        print("Chao titulosCab")

        return 0


